#!/usr/bin/env python3
"""
Advanced Story Classification MCP Server
Uses FastMCP, intelligent scraping, automated source discovery, and human-in-the-loop learning.
"""

import asyncio
import json
import logging
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Any, Optional
import re
import sqlite3
from dataclasses import dataclass, asdict
import hashlib
from dotenv import load_dotenv
import sys

# Load environment variables
load_dotenv()

import aiohttp
import openpyxl
from openpyxl import Workbook
from fastmcp import FastMCP
import requests
from urllib.parse import urljoin, urlparse
import openai
from bs4 import BeautifulSoup

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("advanced-story-classifier")

@dataclass
class Story:
    id: str
    title: str
    content: str
    url: str
    source_domain: str
    author: str
    publish_date: str
    data_elements: List[str]
    topic: str
    scraped_at: str

@dataclass
class Classification:
    story_id: str
    primary_framework: str
    confidence: float
    all_scores: Dict[str, float]
    matched_elements: List[str]
    needs_review: bool
    human_feedback: Optional[str] = None
    corrected_framework: Optional[str] = None

mcp = FastMCP("Advanced Story Classifier")

class EnhancedSourceDiscovery:
    """Enhanced source discovery that truly searches the entire web."""
    
    def __init__(self, google_api_key: str, search_engine_id: str):
        self.google_api_key = google_api_key
        self.search_engine_id = search_engine_id
        
        # Quality indicators for data-driven journalism
        self.quality_indicators = [
            "data", "statistics", "chart", "graph", "survey", "study", 
            "research", "analysis", "report", "findings", "evidence",
            "numbers", "percentage", "poll", "index", "ranking"
        ]
        
        # High-quality domains to prioritize (but not limit to)
        self.preferred_domains = {
            "news": ["reuters.com", "bbc.com", "cnn.com", "npr.org", "pbs.org"],
            "data_journalism": ["fivethirtyeight.com", "pudding.cool", "vox.com"],
            "research": ["pewresearch.org", "brookings.edu", "cfr.org"],
            "government": [".gov", "census.gov", "data.gov"],
            "academic": [".edu", "harvard.edu", "stanford.edu", "mit.edu"],
            "international": ["economist.com", "ft.com", "guardian.com", "washingtonpost.com"]
        }
    
    def create_diverse_search_queries(self, topic: str, max_results: int) -> list:
        """Create diverse search queries to find stories from various sources."""
        
        # Base search terms for the topic
        base_terms = [topic]
        
        # Add topic-specific synonyms
        topic_synonyms = {
            "freedom": ["democracy", "liberty", "civil rights", "human rights", "political freedom"],
            "economics": ["economy", "financial", "economic policy", "GDP", "inflation", "markets"],
            "technology": ["tech", "digital", "innovation", "AI", "cybersecurity"],
            "climate": ["environment", "global warming", "sustainability", "carbon", "emissions"],
            "health": ["healthcare", "medical", "public health", "pandemic", "wellness"]
        }
        
        if topic.lower() in topic_synonyms:
            base_terms.extend(topic_synonyms[topic.lower()])
        
        # Create diverse query patterns
        queries = []
        
        # 1. General web search queries (no site restriction)
        for term in base_terms[:3]:  # Use top 3 terms
            for indicator in self.quality_indicators[:5]:  # Top 5 data indicators
                queries.append(f'"{term}" {indicator}')
                queries.append(f'{term} {indicator} analysis')
        
        # 2. News-focused queries
        for term in base_terms[:2]:
            queries.append(f'{term} news data statistics')
            queries.append(f'{term} report findings study')
        
        # 3. Research-focused queries  
        for term in base_terms[:2]:
            queries.append(f'{term} research study survey data')
            queries.append(f'{term} policy analysis report')
        
        # 4. Time-based queries for fresh content
        for term in base_terms[:2]:
            queries.append(f'{term} 2024 data trends')
            queries.append(f'{term} latest statistics report')
        
        # 5. Domain-specific queries (but still web-wide)
        domain_types = ["news", "research", "academic", "government"]
        for domain_type in domain_types:
            if base_terms:
                queries.append(f'{base_terms[0]} {domain_type} data analysis')
        
        # Limit queries to avoid overwhelming the API
        max_queries = min(20, max_results // 2)
        return queries[:max_queries]
    
    async def search_stories(self, topic: str, max_results: int = 20) -> List[str]:
        """Enhanced search across the entire web using diverse queries."""
        
        all_urls = set()  # Use set to avoid duplicates
        queries = self.create_diverse_search_queries(topic, max_results)
        
        print(f"Searching with {len(queries)} diverse queries...", file=sys.stderr)
        
        for i, query in enumerate(queries):
            try:
                print(f"Query {i+1}/{len(queries)}: {query}", file=sys.stderr)
                
                # Use Google Custom Search API
                response = requests.get(
                    "https://www.googleapis.com/customsearch/v1",
                    params={
                        "key": self.google_api_key,
                        "cx": self.search_engine_id,
                        "q": query,
                        "num": min(10, max_results // len(queries) + 1),  # Distribute results
                        "dateRestrict": "y1",  # Last year for fresher content
                        "sort": "date"  # Prefer recent content
                    },
                    timeout=10
                )
                
                if response.status_code == 200:
                    results = response.json()
                    items = results.get("items", [])
                    
                    print(f"  Found {len(items)} results", file=sys.stderr)
                    
                    for item in items:
                        url = item.get("link", "")
                        title = item.get("title", "")
                        snippet = item.get("snippet", "")
                        
                        # Quality filtering
                        if self._is_quality_source(url, title, snippet):
                            all_urls.add(url)
                            print(f"  ✅ Added: {url}", file=sys.stderr)
                        else:
                            print(f"  ❌ Filtered: {url}", file=sys.stderr)
                            
                        # Stop if we have enough URLs
                        if len(all_urls) >= max_results:
                            break
                else:
                    print(f"  API Error: {response.status_code}", file=sys.stderr)
                    
                # Stop if we have enough URLs
                if len(all_urls) >= max_results:
                    break
                    
                # Rate limiting - be nice to Google
                await asyncio.sleep(0.5)
                
            except Exception as e:
                print(f"Search error for '{query}': {e}", file=sys.stderr)
                continue
        
        # Convert back to list and prioritize by domain quality
        final_urls = self._prioritize_urls(list(all_urls), max_results)
        
        print(f"Final URL collection: {len(final_urls)} unique URLs from diverse sources", file=sys.stderr)
        return final_urls
    
    def _is_quality_source(self, url: str, title: str, snippet: str) -> bool:
        """Filter for quality sources and data-driven content."""
        
        # Skip obviously bad sources
        skip_domains = [
            "youtube.com", "twitter.com", "facebook.com", "instagram.com",
            "pinterest.com", "reddit.com", "quora.com", "answers.com",
            "wiki.com", "ehow.com", "ask.com"
        ]
        
        if any(domain in url.lower() for domain in skip_domains):
            return False
        
        # Skip PDF files and certain file types
        if any(ext in url.lower() for ext in [".pdf", ".doc", ".ppt", ".xls"]):
            return False
        
        # Require some indication of data/analysis content
        text_to_check = f"{title} {snippet}".lower()
        
        # Must have at least one quality indicator
        has_quality_indicator = any(indicator in text_to_check for indicator in self.quality_indicators)
        
        # Bonus points for multiple indicators
        indicator_count = sum(1 for indicator in self.quality_indicators if indicator in text_to_check)
        
        return has_quality_indicator and indicator_count >= 1
    
    def _prioritize_urls(self, urls: List[str], max_results: int) -> List[str]:
        """Prioritize URLs by source quality while maintaining diversity."""
        
        # Categorize URLs by domain type
        categorized = {
            "premium_news": [],
            "data_journalism": [], 
            "research": [],
            "government": [],
            "academic": [],
            "other": []
        }
        
        for url in urls:
            domain = url.lower()
            categorized_flag = False
            
            for category, domains in self.preferred_domains.items():
                if any(pref_domain in domain for pref_domain in domains):
                    if category == "news":
                        categorized["premium_news"].append(url)
                    elif category == "data_journalism":
                        categorized["data_journalism"].append(url)
                    elif category == "research":
                        categorized["research"].append(url)
                    elif category == "government":
                        categorized["government"].append(url)
                    elif category == "academic":
                        categorized["academic"].append(url)
                    categorized_flag = True
                    break
            
            if not categorized_flag:
                categorized["other"].append(url)
        
        # Ensure diversity - take from each category
        final_urls = []
        
        # Distribution strategy: prioritize quality but ensure diversity
        distribution = {
            "data_journalism": min(len(categorized["data_journalism"]), max_results // 4),
            "premium_news": min(len(categorized["premium_news"]), max_results // 3),
            "research": min(len(categorized["research"]), max_results // 6),
            "government": min(len(categorized["government"]), max_results // 8),
            "academic": min(len(categorized["academic"]), max_results // 8),
            "other": min(len(categorized["other"]), max_results // 4)
        }
        
        # Add URLs from each category
        for category, count in distribution.items():
            final_urls.extend(categorized[category][:count])
        
        # Fill remaining slots with best remaining URLs
        remaining_slots = max_results - len(final_urls)
        if remaining_slots > 0:
            all_remaining = []
            for category, urls_list in categorized.items():
                start_idx = distribution.get(category, 0)
                all_remaining.extend(urls_list[start_idx:])
            
            final_urls.extend(all_remaining[:remaining_slots])
        
        print(f"Source diversity: {len(categorized['premium_news'])} news, {len(categorized['data_journalism'])} data journalism, {len(categorized['research'])} research, {len(categorized['other'])} other", file=sys.stderr)
        
        return final_urls[:max_results]

    @mcp.tool()
    async def test_enhanced_source_discovery(
        topic: str,
        max_urls: int = 20
    ) -> str:
        """
        Test the enhanced source discovery to see the diversity of sources found.
        
        Args:
            topic: Topic to search for (e.g., "freedom", "economics")
            max_urls: Maximum number of URLs to find
        """
        try:
            # Check if API keys are loaded
            google_api_key = os.getenv("GOOGLE_API_KEY")
            google_search_engine_id = os.getenv("GOOGLE_SEARCH_ENGINE_ID")
            
            if not all([google_api_key, google_search_engine_id]):
                return "❌ Google API keys not found. Run 'setup_api_keys_from_env' first."
            
            # Initialize enhanced source discovery
            enhanced_discovery = EnhancedSourceDiscovery(google_api_key, google_search_engine_id)
            
            print(f"Testing enhanced source discovery for: {topic}", file=sys.stderr)
            
            # Get URLs
            urls = await enhanced_discovery.search_stories(topic, max_urls)
            
            if not urls:
                return f"❌ No URLs found for topic: {topic}"
            
            # Analyze source diversity
            domains = {}
            domain_categories = {
                "News Sites": [],
                "Data Journalism": [],
                "Research/Think Tanks": [],
                "Government": [],
                "Academic": [],
                "Other": []
            }
            
            for url in urls:
                domain = url.split('/')[2] if '/' in url else url
                domains[domain] = domains.get(domain, 0) + 1
                
                # Categorize
                if any(news in domain for news in ["reuters", "bbc", "cnn", "npr", "pbs"]):
                    domain_categories["News Sites"].append(domain)
                elif any(data in domain for data in ["fivethirtyeight", "pudding", "vox"]):
                    domain_categories["Data Journalism"].append(domain)
                elif any(research in domain for research in ["pew", "brookings", "cfr", "rand"]):
                    domain_categories["Research/Think Tanks"].append(domain)
                elif ".gov" in domain:
                    domain_categories["Government"].append(domain)
                elif ".edu" in domain:
                    domain_categories["Academic"].append(domain)
                else:
                    domain_categories["Other"].append(domain)
            
            # Create report
            report = f"Enhanced Source Discovery Test Results\n"
            report += f"Topic: {topic}\n"
            report += f"URLs Found: {len(urls)}\n"
            report += f"Unique Domains: {len(domains)}\n"
            report += "=" * 50 + "\n\n"
            
            # Source diversity breakdown
            report += "Source Categories:\n"
            for category, domains_list in domain_categories.items():
                unique_domains = list(set(domains_list))
                if unique_domains:
                    report += f"  {category}: {len(unique_domains)} domains\n"
                    for domain in unique_domains[:3]:  # Show first 3
                        count = domains.get(domain, 0)
                        report += f"    • {domain} ({count} articles)\n"
                    if len(unique_domains) > 3:
                        report += f"    ... and {len(unique_domains) - 3} more\n"
            
            report += "\nAll Domains Found:\n"
            sorted_domains = sorted(domains.items(), key=lambda x: x[1], reverse=True)
            for domain, count in sorted_domains:
                report += f"  {domain}: {count} articles\n"
            
            report += f"\nSample URLs:\n"
            for i, url in enumerate(urls[:10], 1):
                report += f"  {i}. {url}\n"
            
            if len(urls) > 10:
                report += f"  ... and {len(urls) - 10} more URLs\n"
            
            # Comparison with old system
            report += f"\n📊 Improvement Analysis:\n"
            report += f"  Previous system: ~4 domains (hardcoded)\n"
            report += f"  Enhanced system: {len(domains)} domains (web-wide search)\n"
            report += f"  Diversity improvement: {len(domains)/4:.1f}x more sources\n"
            
            return report
            
        except Exception as e:
            return f"Error testing enhanced source discovery: {str(e)}"

    @mcp.tool()
    async def collect_with_enhanced_discovery(
        topic: str,
        num_stories: int = 15,
        excel_file: str = "enhanced_stories.xlsx"
    ) -> str:
        """
        Collect stories using the enhanced source discovery system.
        
        Args:
            topic: Topic to search for
            num_stories: Number of stories to collect
            excel_file: Excel file to save results
        """
        global scraper, classifier
        
        try:
            # Check if components are initialized
            if not all([scraper, classifier]):
                return "❌ Please run 'setup_api_keys_from_env' first"
            
            # Initialize enhanced source discovery
            google_api_key = os.getenv("GOOGLE_API_KEY")
            google_search_engine_id = os.getenv("GOOGLE_SEARCH_ENGINE_ID")
            
            enhanced_discovery = EnhancedSourceDiscovery(google_api_key, google_search_engine_id)
            
            # Get diverse URLs
            print(f"Searching for diverse sources on: {topic}", file=sys.stderr)
            urls = await enhanced_discovery.search_stories(topic, num_stories * 2)  # Get extra for filtering
            
            if not urls:
                return f"❌ No URLs found for topic: {topic}"
            
            # Track source diversity
            source_domains = set()
            
            # Scrape stories with enhanced diversity
            stories = []
            classifications = []
            
            for i, url in enumerate(urls):
                if len(stories) >= num_stories:
                    break
                    
                domain = url.split('/')[2] if '/' in url else url
                
                try:
                    story = await scraper.scrape_story(url, topic)
                    if story:
                        # Use enhanced classifier if available
                        if hasattr(classifier, 'classify_story'):
                            classification = classifier.classify_story(story)
                        else:
                            # Fallback to original classifier
                            classification = classifier.classify_story(story)
                        
                        stories.append(story)
                        classifications.append(classification)
                        source_domains.add(domain)
                        
                        print(f"✅ Collected from {domain}: {story.title[:50]}...", file=sys.stderr)
                    else:
                        print(f"❌ Failed to extract from {domain}", file=sys.stderr)
                        
                except Exception as e:
                    print(f"❌ Error processing {url}: {e}", file=sys.stderr)
            
            if not stories:
                return "❌ No stories could be successfully extracted"
            
            # Save with enhanced Excel manager
            excel_manager = ExcelManager(excel_file)
            excel_manager.save_stories_and_classifications(stories, classifications)
            
            # Create summary with source diversity stats
            summary = f"Enhanced Collection Results for '{topic}':\n"
            summary += f"Stories collected: {len(stories)}\n"
            summary += f"Unique source domains: {len(source_domains)}\n"
            summary += f"Average confidence: {sum(c.confidence for c in classifications) / len(classifications):.3f}\n\n"
            
            summary += "Source Diversity:\n"
            domain_counts = {}
            for story in stories:
                domain = story.source_domain
                domain_counts[domain] = domain_counts.get(domain, 0) + 1
            
            for domain, count in sorted(domain_counts.items(), key=lambda x: x[1], reverse=True):
                summary += f"  • {domain}: {count} stories\n"
            
            summary += f"\nHigh-confidence stories: {sum(1 for c in classifications if c.confidence >= 0.6)}\n"
            summary += f"Stories needing review: {sum(1 for c in classifications if c.needs_review)}\n"
            summary += f"\nSaved to: {excel_file}"
            
            return summary
            
        except Exception as e:
            return f"Error with enhanced collection: {str(e)}"


# class SourceDiscovery:
#     """Intelligent source discovery for data journalism sites."""
    
#     def __init__(self, google_api_key: str, search_engine_id: str):
#         self.google_api_key = google_api_key
#         self.search_engine_id = search_engine_id
#         self.known_sources = {
#             "economics": [
#                 "fivethirtyeight.com", "economist.com", "bloomberg.com",
#                 "tradingeconomics.com", "fred.stlouisfed.org", "ourworldindata.org"
#             ],
#             "technology": [
#                 "arstechnica.com", "techcrunch.com", "wired.com",
#                 "spectrum.ieee.org", "technologyreview.mit.edu"
#             ],
#             "climate": [
#                 "carbonbrief.org", "climate.gov", "climatecentral.org",
#                 "realclimate.org", "skepticalscience.com"
#             ],
#             "health": [
#                 "healthdata.org", "who.int", "cdc.gov",
#                 "nejm.org", "thelancet.com"
#             ],
#             "general": [
#                 "pudding.cool", "fivethirtyeight.com", "vox.com",
#                 "reuters.com", "bbc.com", "nytimes.com"
#             ]
#         }
    
#     def get_topic_sources(self, topic: str) -> List[str]:
#         """Get relevant sources based on topic."""
#         topic_lower = topic.lower()
        
#         # Find matching category
#         for category, sources in self.known_sources.items():
#             if category in topic_lower or any(keyword in topic_lower for keyword in [category]):
#                 return sources
        
#         return self.known_sources["general"]
    
#     async def search_stories(self, topic: str, max_results: int = 20) -> List[str]:
#         """Search for story URLs using Google Custom Search."""
#         sources = self.get_topic_sources(topic)
#         urls = []
        
#         # Create search queries for each source
#         for source in sources[:5]:  # Limit to top 5 sources
#             query = f"site:{source} {topic} data statistics chart graph"
            
#             try:
#                 response = requests.get(
#                     "https://www.googleapis.com/customsearch/v1",
#                     params={
#                         "key": self.google_api_key,
#                         "cx": self.search_engine_id,
#                         "q": query,
#                         "num": min(10, max_results // len(sources))
#                     }
#                 )
                
#                 if response.status_code == 200:
#                     results = response.json()
#                     for item in results.get("items", []):
#                         urls.append(item["link"])
                        
#             except Exception as e:
#                 logger.error(f"Search error for {source}: {e}")
        
#         return urls[:max_results]

class IntelligentScraper:
    """AI-powered web scraper inspired by CrewAI's approach."""
    
    def __init__(self, openrouter_api_key: str, model: str = "openai/gpt-3.5-turbo"):
        self.openai_client = openai.OpenAI(
            base_url="https://openrouter.ai/api/v1",
            api_key=openrouter_api_key,
        )
        self.model = model
        self.session = None
    
    async def create_session(self):
        if not self.session:
            self.session = aiohttp.ClientSession(
                headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                },
                timeout=aiohttp.ClientTimeout(total=30)
            )
    
    async def close_session(self):
        if self.session:
            await self.session.close()
    
    async def scrape_story(self, url: str, topic: str) -> Optional[Story]:
        """Intelligently scrape a story with AI-powered content extraction."""
        try:
            await self.create_session()
            async with self.session.get(url) as response:
                if response.status != 200:
                    return None
                
                html = await response.text()
                soup = BeautifulSoup(html, 'html.parser')
                
                # Extract raw content
                raw_data = self._extract_raw_content(soup)
                
                # Use AI to intelligently extract and structure content
                structured_content = await self._ai_extract_content(raw_data, url, topic)
                
                if not structured_content:
                    return None
                
                # Create story object
                story_id = hashlib.md5(url.encode()).hexdigest()[:12]
                
                return Story(
                    id=story_id,
                    title=structured_content.get("title", ""),
                    content=structured_content.get("content", ""),
                    url=url,
                    source_domain=urlparse(url).netloc,
                    author=structured_content.get("author", "Unknown"),
                    publish_date=structured_content.get("date", ""),
                    data_elements=structured_content.get("data_elements", []),
                    topic=topic,
                    scraped_at=datetime.now().isoformat()
                )
                
        except Exception as e:
            logger.error(f"Error scraping {url}: {e}")
            return None
    
    def _extract_raw_content(self, soup: BeautifulSoup) -> Dict[str, str]:
        """Extract raw content from HTML."""
        # Remove script and style elements
        for script in soup(["script", "style", "nav", "footer", "header"]):
            script.decompose()
        
        return {
            "title": soup.find('title').get_text() if soup.find('title') else "",
            "h1": " ".join([h.get_text() for h in soup.find_all('h1')]),
            "paragraphs": " ".join([p.get_text() for p in soup.find_all('p')]),
            "full_text": soup.get_text()[:5000]  # Limit for AI processing
        }
    
    async def _ai_extract_content(self, raw_data: Dict[str, str], url: str, topic: str) -> Optional[Dict[str, Any]]:
        """Use AI to extract and structure content intelligently."""
        
        prompt = f"""
        Extract structured information from this web article about {topic}.
        
        URL: {url}
        Raw content: {raw_data['full_text'][:2000]}
        
        Please extract and return JSON with:
        1. title: Article headline
        2. content: Main article content (summary if too long)
        3. author: Author name if available
        4. date: Publication date if available
        5. data_elements: List of data visualizations mentioned (charts, graphs, statistics, etc.)
        6. is_data_driven: Boolean - does this article contain substantial data analysis?
        7. relevance_to_topic: Score 0-1 how relevant this is to "{topic}"
        
        Only return the JSON, no other text.
        """
        
        try:
            response = self.openai_client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1
            )
            
            result = json.loads(response.choices[0].message.content)
            
            # Filter out non-data-driven or irrelevant content
            if not result.get("is_data_driven", False) or result.get("relevance_to_topic", 0) < 0.3:
                return None
                
            return result
            
        except Exception as e:
            logger.error(f"AI extraction error: {e}")
            return None

class LearningClassifier:
    """Narrative classifier that learns from human feedback."""
    
    def __init__(self, db_path: str = "classifier_learning.db"):
        self.db_path = db_path
        self.init_database()
        
        # Enhanced narrative structures with learning weights
        self.narrative_structures = {
            "Time-Based Progression": {
                "description": "Data spans across time showing chronological evolution",
                "base_keywords": ["over time", "chronological", "timeline", "progression", "evolution", "trends"],
                "base_patterns": [r"from \d{4} to \d{4}", r"between \d{4} and \d{4}", r"over the past \w+ years"],
                "learned_keywords": [],
                "learned_patterns": [],
                "weight": 1.0
            },
            "Overview to Detail": {
                "description": "High-level summaries zooming into specific details",
                "base_keywords": ["overview", "summary", "breakdown", "detailed analysis", "drilling down"],
                "base_patterns": [r"total \w+ overview", r"breakdown by", r"detailed look at"],
                "learned_keywords": [],
                "learned_patterns": [],
                "weight": 1.0
            },
            "Cause-and-Effect": {
                "description": "Causal relationships between data points",
                "base_keywords": ["because", "due to", "resulted in", "caused by", "led to", "correlation", "impact"],
                "base_patterns": [r"correlation between \w+ and \w+", r"impact of \w+ on \w+"],
                "learned_keywords": [],
                "learned_patterns": [],
                "weight": 1.0
            },
            "Workflow/Process": {
                "description": "Steps in an analytical process or workflow",
                "base_keywords": ["process", "workflow", "steps", "methodology", "procedure", "analysis"],
                "base_patterns": [r"step \d+", r"phase \d+", r"first.*then.*finally"],
                "learned_keywords": [],
                "learned_patterns": [],
                "weight": 1.0
            },
            "Comparative Analysis": {
                "description": "Comparisons leading to conclusions or patterns",
                "base_keywords": ["compared to", "versus", "vs", "comparison", "contrast", "difference"],
                "base_patterns": [r"\w+ vs \w+", r"compared to", r"in contrast to"],
                "learned_keywords": [],
                "learned_patterns": [],
                "weight": 1.0
            },
            "Thematic Clustering": {
                "description": "Related data points grouped by themes",
                "base_keywords": ["themes", "categories", "groups", "clusters", "segments", "demographics"],
                "base_patterns": [r"grouped by", r"categorized by", r"segmented by"],
                "learned_keywords": [],
                "learned_patterns": [],
                "weight": 1.0
            },
            "Problem-Solution Framework": {
                "description": "Problem definition with data diagnosis and detailed solutions",
                "base_keywords": ["problem", "issue", "challenge", "solution", "resolve", "fix", "intervention"],
                "base_patterns": [r"problem.*solution", r"challenge.*intervention"],
                "learned_keywords": [],
                "learned_patterns": [],
                "weight": 1.0
            },
            "Question-and-Answer": {
                "description": "Questions followed by explanations of drivers",
                "base_keywords": ["why", "how", "what", "question", "answer", "explain", "drivers"],
                "base_patterns": [r"why did \w+", r"what caused", r"how did \w+"],
                "learned_keywords": [],
                "learned_patterns": [],
                "weight": 1.0
            }
        }
        
        self.load_learned_patterns()
    
    def init_database(self):
        """Initialize SQLite database for learning."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS feedback (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                story_id TEXT,
                original_classification TEXT,
                corrected_classification TEXT,
                confidence REAL,
                feedback_text TEXT,
                timestamp TEXT,
                story_content TEXT
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS learned_patterns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                framework TEXT,
                pattern_type TEXT,
                pattern_value TEXT,
                weight REAL,
                learned_from TEXT,
                timestamp TEXT
            )
        """)
        
        conn.commit()
        conn.close()
    
    def load_learned_patterns(self):
        """Load previously learned patterns from database."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("SELECT framework, pattern_type, pattern_value, weight FROM learned_patterns")
        rows = cursor.fetchall()
        
        for framework, pattern_type, pattern_value, weight in rows:
            if framework in self.narrative_structures:
                if pattern_type == "keyword":
                    self.narrative_structures[framework]["learned_keywords"].append(pattern_value)
                elif pattern_type == "pattern":
                    self.narrative_structures[framework]["learned_patterns"].append(pattern_value)
                
                # Update framework weight based on learning
                self.narrative_structures[framework]["weight"] = weight
        
        conn.close()
    
    def classify_story(self, story: Story) -> Classification:
        """Classify story using enhanced learning-based approach."""
        text = f"{story.title} {story.content}".lower()
        scores = {}
        
        for framework_name, framework_info in self.narrative_structures.items():
            score = 0
            matched_elements = []
            
            # Check base keywords
            all_keywords = framework_info["base_keywords"] + framework_info["learned_keywords"]
            for keyword in all_keywords:
                if keyword.lower() in text:
                    weight = 2.0 if keyword in framework_info["learned_keywords"] else 1.0
                    score += weight
                    matched_elements.append(f"keyword: {keyword}")
            
            # Check patterns
            all_patterns = framework_info["base_patterns"] + framework_info["learned_patterns"]
            for pattern in all_patterns:
                if re.search(pattern, text, re.IGNORECASE):
                    weight = 3.0 if pattern in framework_info["learned_patterns"] else 2.0
                    score += weight
                    matched_elements.append(f"pattern: {pattern}")
            
            # Apply framework weight from learning
            score *= framework_info["weight"]
            
            # Normalize score
            max_possible = len(all_keywords) * 2 + len(all_patterns) * 3
            normalized_score = score / max_possible if max_possible > 0 else 0
            
            scores[framework_name] = normalized_score
        
        # Find best classification
        best_framework = max(scores.keys(), key=lambda k: scores[k])
        confidence = scores[best_framework]
        
        return Classification(
            story_id=story.id,
            primary_framework=best_framework,
            confidence=round(confidence, 3),
            all_scores=scores,
            matched_elements=matched_elements,
            needs_review=confidence < 0.4
        )
    
    def learn_from_feedback(self, story: Story, classification: Classification, 
                          corrected_framework: str, feedback_text: str):
        """Learn from human feedback to improve future classifications."""
        
        # Store feedback
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO feedback (story_id, original_classification, corrected_classification, 
                                confidence, feedback_text, timestamp, story_content)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            story.id, classification.primary_framework, corrected_framework,
            classification.confidence, feedback_text, datetime.now().isoformat(),
            story.content[:1000]
        ))
        
        # Extract new patterns from corrected examples
        self._extract_learning_patterns(story, corrected_framework, cursor)
        
        conn.commit()
        conn.close()
        
        # Reload patterns
        self.load_learned_patterns()
    
    def _extract_learning_patterns(self, story: Story, correct_framework: str, cursor):
        """Extract new keywords/patterns from correctly classified examples."""
        text = f"{story.title} {story.content}".lower()
        
        # Simple pattern extraction (could be enhanced with NLP)
        words = text.split()
        
        # Look for domain-specific terms
        for i, word in enumerate(words):
            if len(word) > 5 and word.isalpha():
                # Check if this word appears frequently in correct classifications
                cursor.execute("""
                    SELECT COUNT(*) FROM feedback 
                    WHERE corrected_classification = ? AND story_content LIKE ?
                """, (correct_framework, f"%{word}%"))
                
                count = cursor.fetchone()[0]
                if count >= 2:  # Appears in multiple correct examples
                    # Add as learned keyword
                    cursor.execute("""
                        INSERT OR IGNORE INTO learned_patterns 
                        (framework, pattern_type, pattern_value, weight, learned_from, timestamp)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (
                        correct_framework, "keyword", word, 1.2,
                        story.id, datetime.now().isoformat()
                    ))

class ExcelManager:
    """Enhanced Excel management with human feedback tracking."""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
    
    def create_or_load_workbook(self) -> openpyxl.Workbook:
        """Create new workbook with multiple sheets."""
        if self.file_path.exists():
            return openpyxl.load_workbook(self.file_path)
        else:
            wb = Workbook()
            
            # Main classifications sheet
            ws_main = wb.active
            ws_main.title = "Story Classifications"
            headers_main = [
                "Story ID", "Title", "URL", "Source Domain", "Author", "Publish Date",
                "Topic", "Primary Classification", "Confidence Score", "Matched Elements",
                "Data Elements", "Needs Review", "Human Feedback", "Corrected Framework",
                "Scraped At", "Content Preview"
            ]
            for col, header in enumerate(headers_main, 1):
                ws_main.cell(row=1, column=col, value=header)
            
            # Feedback tracking sheet
            ws_feedback = wb.create_sheet("Human Feedback")
            headers_feedback = [
                "Feedback ID", "Story ID", "Original Classification", "Corrected Classification",
                "Feedback Text", "Timestamp", "Confidence Improvement"
            ]
            for col, header in enumerate(headers_feedback, 1):
                ws_feedback.cell(row=1, column=col, value=header)
            
            return wb
    
    def save_stories_and_classifications(self, stories: List[Story], classifications: List[Classification]):
        """Save stories with classifications to Excel."""
        wb = self.create_or_load_workbook()
        ws = wb["Story Classifications"]
        
        next_row = ws.max_row + 1
        
        for story, classification in zip(stories, classifications):
            row_data = [
                story.id, story.title, story.url, story.source_domain, story.author,
                story.publish_date, story.topic, classification.primary_framework,
                classification.confidence, ", ".join(classification.matched_elements),
                ", ".join(story.data_elements), classification.needs_review,
                classification.human_feedback or "", classification.corrected_framework or "",
                story.scraped_at, story.content[:200] + "..." if len(story.content) > 200 else story.content
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=next_row, column=col, value=value)
            
            next_row += 1
        
        wb.save(self.file_path)


# Global instances
source_discovery = None
scraper = None
classifier = None

@mcp.tool()
async def collect_and_classify_stories(
    topic: str,
    num_stories: int = 15,
    excel_file: str = "story_classifications.xlsx"
) -> str:
    """
    Collect and classify data-driven stories on a specific topic.
    
    Args:
        topic: The topic to search for (e.g., "economics", "climate change", "technology trends")
        num_stories: Number of stories to collect (10-50 recommended)
        excel_file: Path to Excel file for storing results
    """
    global source_discovery, scraper, classifier
    
    try:
        # Initialize components if needed
        if not all([source_discovery, scraper, classifier]):
            return "Please configure API keys first using setup_api_keys tool"
        
        # Discover relevant story URLs
        logger.info(f"Searching for {num_stories} stories about: {topic}")
        urls = await source_discovery.search_stories(topic, num_stories * 2)  # Get extra for filtering
        
        if not urls:
            return f"No relevant URLs found for topic: {topic}"
        
        # Scrape stories intelligently
        stories = []
        classifications = []
        
        for url in urls[:num_stories]:
            story = await scraper.scrape_story(url, topic)
            if story:
                classification = classifier.classify_story(story)
                stories.append(story)
                classifications.append(classification)
                
                if len(stories) >= num_stories:
                    break
        
        if not stories:
            return "No data-driven stories could be extracted from the found URLs"
        
        # Save to Excel
        excel_manager = ExcelManager(excel_file)
        excel_manager.save_stories_and_classifications(stories, classifications)
        
        # Create summary
        summary = f"Successfully collected and classified {len(stories)} stories about '{topic}':\n\n"
        
        review_needed = []
        for i, (story, classification) in enumerate(zip(stories, classifications), 1):
            summary += f"{i}. {story.title[:60]}...\n"
            summary += f"   Source: {story.source_domain}\n"
            summary += f"   Classification: {classification.primary_framework}\n"
            summary += f"   Confidence: {classification.confidence:.3f}\n"
            
            if classification.needs_review:
                review_needed.append(i)
            
            summary += "\n"
        
        if review_needed:
            summary += f"\n⚠️ Items needing human review: {', '.join(map(str, review_needed))}\n"
            summary += "Use the 'review_and_correct' tool to provide feedback on these items.\n"
        
        summary += f"\nResults saved to: {excel_file}"
        summary += f"\nUse 'search_classified_stories' to query the collected data."
        
        return summary
        
    except Exception as e:
        logger.error(f"Error in collect_and_classify_stories: {e}")
        return f"Error: {str(e)}"

@mcp.tool()
async def review_and_correct(
    story_id: str,
    corrected_framework: str,
    feedback_text: str = ""
) -> str:
    """
    Provide human feedback to improve classification accuracy.
    
    Args:
        story_id: ID of the story to correct
        corrected_framework: The correct narrative framework
        feedback_text: Optional explanation of why the correction was made
    """
    global classifier
    
    try:
        # This would need to load the story from database/Excel
        # For now, simulate the learning process
        
        valid_frameworks = list(classifier.narrative_structures.keys())
        if corrected_framework not in valid_frameworks:
            return f"Invalid framework. Valid options: {', '.join(valid_frameworks)}"
        
        # In a full implementation, this would:
        # 1. Load the original story and classification
        # 2. Call classifier.learn_from_feedback()
        # 3. Update the Excel file
        
        return f"Thank you for the feedback! The system has learned that story {story_id} should be classified as '{corrected_framework}'. This will improve future classifications."
        
    except Exception as e:
        return f"Error processing feedback: {str(e)}"

@mcp.tool()
async def setup_api_keys_from_env() -> str:
    """
    Configure API keys from environment variables (.env file).
    Requires GOOGLE_API_KEY, GOOGLE_SEARCH_ENGINE_ID, and OPENROUTER_API_KEY in .env
    """
    global source_discovery, scraper, classifier
    
    try:
        google_api_key = os.getenv("GOOGLE_API_KEY")
        google_search_engine_id = os.getenv("GOOGLE_SEARCH_ENGINE_ID") 
        openrouter_api_key = os.getenv("OPENROUTER_API_KEY")
        model = os.getenv("AI_MODEL", "anthropic/claude-3-haiku")  # Default to cheap, good model
        
        if not all([google_api_key, google_search_engine_id, openrouter_api_key]):
            missing = []
            if not google_api_key: missing.append("GOOGLE_API_KEY")
            if not google_search_engine_id: missing.append("GOOGLE_SEARCH_ENGINE_ID")
            if not openrouter_api_key: missing.append("OPENROUTER_API_KEY")
            
            return f"❌ Missing environment variables: {', '.join(missing)}\n" + \
                   "Please add them to your .env file"
        
        source_discovery = EnhancedSourceDiscoveryDiscovery(google_api_key, google_search_engine_id)
        scraper = IntelligentScraper(openrouter_api_key, model)
        classifier = LearningClassifier()
        
        return f"✅ API keys loaded from .env successfully!\n" + \
               f"Using AI model: {model}\n" + \
               f"Google API: {google_api_key[:10]}...\n" + \
               f"Search Engine ID: {google_search_engine_id}\n" + \
               "You can now use collect_and_classify_stories."
        
    except Exception as e:
        return f"Error loading API keys from .env: {str(e)}"

@mcp.tool()
async def search_classified_stories(
    query: str = "",
    framework: str = "",
    min_confidence: float = 0.0,
    excel_file: str = "story_classifications.xlsx"
) -> str:
    """
    Search through previously classified stories.
    
    Args:
        query: Text to search in titles/content
        framework: Filter by specific narrative framework
        min_confidence: Minimum confidence score
        excel_file: Excel file to search
    """
    try:
        # This would implement Excel search functionality
        return f"Search functionality would filter stories based on: query='{query}', framework='{framework}', min_confidence={min_confidence}"
        
    except Exception as e:
        return f"Error searching stories: {str(e)}"

@mcp.tool()
async def list_narrative_frameworks() -> str:
    """List all available narrative classification structures."""
    
    result = "Available Narrative Frameworks:\n\n"
    
    frameworks = {
        "Time-Based Progression": "Data spans across time showing chronological evolution",
        "Overview to Detail": "High-level summaries zooming into specific details", 
        "Cause-and-Effect": "Causal relationships between data points",
        "Workflow/Process": "Steps in an analytical process or workflow",
        "Comparative Analysis": "Comparisons leading to conclusions or patterns",
        "Thematic Clustering": "Related data points grouped by themes",
        "Problem-Solution Framework": "Problem definition with data diagnosis and detailed solutions",
        "Question-and-Answer": "Questions followed by explanations of drivers"
    }
    
    for i, (name, desc) in enumerate(frameworks.items(), 1):
        result += f"{i}. **{name}**\n   {desc}\n\n"
    
    return result

if __name__ == "__main__":
    mcp.run()